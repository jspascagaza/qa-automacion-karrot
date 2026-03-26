from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os

import argparse

import random

def generate_data(
    filename,
    product_count,
    variants_per_product=2,
    products_with_variants=None,
    attribute_count=2,
    variants_per_product_mix=None,
    material_blank_rate=0.0,
    color_options=None,
    duplicate_sku_count=0,
    category_options=None,
    inject_errors=False,
    error_rate=0.05,
    seed=None,
):
    file_path = os.path.join(os.getcwd(), filename)

    if int(product_count) <= 0:
        raise ValueError("product_count must be >= 1")

    if int(variants_per_product) <= 0:
        raise ValueError("variants_per_product must be >= 1")

    if products_with_variants is None:
        products_with_variants = min(2, int(product_count))

    if int(products_with_variants) < 0:
        raise ValueError("products_with_variants must be >= 0")

    if int(products_with_variants) > int(product_count):
        raise ValueError("products_with_variants must be <= product_count")

    if int(attribute_count) not in (1, 2):
        raise ValueError("attribute_count must be 1 or 2")

    if variants_per_product_mix is not None:
        if not isinstance(variants_per_product_mix, (list, tuple)) or not variants_per_product_mix:
            raise ValueError("variants_per_product_mix must be a non-empty list")
        mix = []
        for v in variants_per_product_mix:
            iv = int(v)
            if iv <= 0:
                raise ValueError("variants_per_product_mix values must be >= 1")
            mix.append(iv)
        variants_per_product_mix = mix

    if float(material_blank_rate) < 0 or float(material_blank_rate) > 1:
        raise ValueError("material_blank_rate must be between 0 and 1")

    if color_options is not None:
        if not isinstance(color_options, (list, tuple)) or not color_options:
            raise ValueError("color_options must be a non-empty list")
        color_options = [str(c).strip() for c in color_options if str(c).strip()]
        if not color_options:
            raise ValueError("color_options must include at least one non-empty value")

    if int(duplicate_sku_count) < 0:
        raise ValueError("duplicate_sku_count must be >= 0")

    if category_options is not None:
        if not isinstance(category_options, (list, tuple)) or not category_options:
            raise ValueError("category_options must be a non-empty list")
        category_options = [str(c).strip() for c in category_options if str(c).strip()]
        if not category_options:
            raise ValueError("category_options must include at least one non-empty value")

    if seed is not None:
        random.seed(seed)

    products_with_variants_count = int(products_with_variants)
    simple_products_count = int(product_count) - products_with_variants_count

    if variants_per_product_mix is not None:
        variants_count_by_product = [random.choice(variants_per_product_mix) for _ in range(products_with_variants_count)]
    else:
        variants_count_by_product = [int(variants_per_product)] * products_with_variants_count

    total_rows = sum(variants_count_by_product) + simple_products_count
    print(
        f"Starting generation of {filename} with {total_rows} rows "
        f"({products_with_variants_count} products with variants, {simple_products_count} simple products)..."
    )

    # Define Styles
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    pink_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    grey_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    header_font = Font(bold=True, size=11)
    
    # Write-only workbook (streaming)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Plantilla Productos")

    # Column headers and widths
    columns = [
        ('Type*', 15),
        ('Name*', 30),
        ('Fotografía producto', 25),
        ('Category*', 25),
        ('Grupo de Unidades', 25),
        ('Unit*', 15),
        ('Description*', 40),
        ('isPerishable*', 15),
        ('Atributos', 15),
        ('Atributos', 15),
        ('SKU*', 20),
        ('Barcode*', 20),
        ('Barcode*', 15),
        ('Barcode*', 15),
        ('Cost*', 15),
        ('Price*', 15),
        ('PUM', 15),
        ('PUM', 15),
        ('Proveedor', 20),
        ('Proveedor', 15),
    ]

    # Set column widths (approximate for WriteOnly)
    # Note: in write_only, we can't easily set column dimensions upfront for all columns in a standard way 
    # without accessing sheet properties, but we'll focus on cell content first.
    
    # Helper to create styled cell
    def create_styled_cell(value, fill=None, font=None, border=None, alignment=None):
        cell = WriteOnlyCell(ws, value=value)
        if fill: cell.fill = fill
        if font: cell.font = font
        if border: cell.border = border
        if alignment: cell.alignment = alignment
        return cell

    # Row 1: Main headers (no placeholder rows - frontend expects headers in row 1)
    # Headers must match what processCsvData.js expects (keyMapping aliases)
    final_headers_row_5 = [
        'Type*', 'Name*', 'Fotografía producto', 'Category*',
        'Grupo de Unidades', 'Unit*', 'Description*',
        'isPerishable*',
        'Attribute color',
        *(['Attribute material'] if int(attribute_count) == 2 else []),
        'SKU',                  # SKU without asterisk (matches keyMapping)
        'Barcode',              # Primary barcode
        'Código 2',             # BarcodeTwo alias
        'Código 3',             # BarcodeThree alias
        'Cost',                 # Cost without asterisk
        'Price*',               # Price with asterisk
        'Base (Unidad)',        # PUMUnit alias
        'Valor',                # PUMValue alias  
        'Proveedor',            # VendorID alias (Nombre)
        'Costo Proveedor'       # VendorCost alias
    ]

    row5_cells = []
    for h in final_headers_row_5:
        row5_cells.append(create_styled_cell(h, fill=header_fill, font=header_font, border=thin_border, alignment=Alignment(horizontal='center')))
    ws.append(row5_cells)

    # No sub-headers row - data starts immediately after headers

    # Unit Configuration from User
    units_config = [
        {
            "companyID": "c09c3c41-27ae-4906-a49a-292f7c16622d",
            "createdAt": "2025-10-21T14:25:00.000Z",
            "description": "Unidades de medida para volumen o capacidad",
            "id": "e5a94b7e-4f5a-4c60-b517-273b36cb1d5a",
            "name": "Volumen",
            "updatedAt": "2025-10-30T18:17:51.564Z",
            "units": [
                {"abbr": "pt", "name": "Pinta (EE. UU.)"},
                {"abbr": "gal", "name": "Galón (EE. UU.)"},
                {"abbr": "m³", "name": "Metro cúbico"},
                {"abbr": "mL", "name": "Mililitro"},
                {"abbr": "fl oz", "name": "Onza líquida (EE. UU.)"},
                {"abbr": "L", "name": "Litro"}
            ]
        },
        {
            "companyID": "c09c3c41-27ae-4906-a49a-292f7c16622d",
            "createdAt": "2025-10-21T14:30:00.000Z",
            "description": "Unidades de medida para superficie o extensión",
            "id": "fc7f0b41-6c35-4b29-bd77-8ac5de38b926",
            "name": "Área",
            "updatedAt": "2025-10-30T18:17:51.564Z",
            "units": [
                {"abbr": "ac", "name": "Acre"},
                {"abbr": "yd²", "name": "Yarda cuadrada"},
                {"abbr": "km²", "name": "Kilómetro cuadrado"},
                {"abbr": "ft²", "name": "Pie cuadrado"},
                {"abbr": "mm²", "name": "Milímetro cuadrado"},
                {"abbr": "cm²", "name": "Centímetro cuadrado"},
                {"abbr": "ha", "name": "Hectárea"},
                {"abbr": "m²", "name": "Metro cuadrado"}
            ]
        },
        {
            "companyID": "c09c3c41-27ae-4906-a49a-292f7c16622d",
            "createdAt": "2025-10-21T14:40:00.000Z",
            "description": "Unidades de medida para conteo o cantidad de objetos",
            "id": "f7c82967-5cb1-42b8-9213-792d8cf5e3f4",
            "name": "Cantidad / Unidades",
            "updatedAt": "2025-10-30T18:17:51.564Z",
            "units": [
                {"abbr": "dz", "name": "Docena"},
                {"abbr": "c", "name": "Ciento"},
                {"abbr": "u", "name": "Unidad"},
                {"abbr": "k", "name": "Millar"}
            ]
        },
        {
            "companyID": "c09c3c41-27ae-4906-a49a-292f7c16622d",
            "createdAt": "2025-10-21T14:15:00.000Z",
            "description": "Unidades de medida para peso o masa",
            "id": "7b0f1d34-56f4-47c8-8a3e-1c59f1f95f1b",
            "name": "Peso / Masa",
            "updatedAt": "2025-10-30T18:17:51.564Z",
            "units": [
                {"abbr": "t", "name": "Tonelada"},
                {"abbr": "lb", "name": "Libra"},
                {"abbr": "mg", "name": "Miligramo"},
                {"abbr": "qq", "name": "Quintal"},
                {"abbr": "kg", "name": "Kilogramo"},
                {"abbr": "oz", "name": "Onza"},
                {"abbr": "g", "name": "Gramo"}
            ]
        },
        {
            "companyID": "c09c3c41-27ae-4906-a49a-292f7c16622d",
            "createdAt": "2025-10-21T14:00:00.000Z",
            "description": "Unidades de medida para longitud o distancia",
            "id": "9e8b6a56-7a0b-4b8a-b8f2-77d6b6b60c7e",
            "name": "Distancia / Longitud",
            "updatedAt": "2025-10-30T18:17:51.564Z",
            "units": [
                {"abbr": "µm", "name": "Micrómetro"},
                {"abbr": "m", "name": "Metro"},
                {"abbr": "mi", "name": "Milla"},
                {"abbr": "km", "name": "Kilómetro"},
                {"abbr": "mm", "name": "Milímetro"},
                {"abbr": "ft", "name": "Pie"},
                {"abbr": "in", "name": "Pulgada"},
                {"abbr": "cm", "name": "Centímetro"},
                {"abbr": "yd", "name": "Yarda"}
            ]
        }
    ]

    # Tech Configurations - Expanded for unique combinations
    tech_products = {
        "Dispositivos móviles y tecnología": [
            ("iPhone 13", [64, 128, 256, 512], "GB", 6000000),
            ("iPhone 13 Mini", [64, 128, 256, 512], "GB", 5500000),
            ("iPhone 13 Pro", [128, 256, 512, 1024], "GB", 7500000),
            ("iPhone 13 Pro Max", [128, 256, 512, 1024], "GB", 8000000),
            ("iPhone 14", [128, 256, 512], "GB", 7000000),
            ("iPhone 14 Plus", [128, 256, 512], "GB", 7500000),
            ("iPhone 14 Pro", [128, 256, 512, 1024], "GB", 8000000),
            ("iPhone 14 Pro Max", [256, 512, 1024], "GB", 8500000),
            ("iPhone 15", [128, 256, 512], "GB", 7500000),
            ("iPhone 15 Plus", [128, 256, 512], "GB", 8000000),
            ("iPhone 15 Pro", [128, 256, 512, 1024], "GB", 8500000),
            ("iPhone 15 Pro Max", [256, 512, 1024], "GB", 9000000),
            ("iPhone SE 2022", [64, 128, 256], "GB", 2500000),
            ("Samsung Galaxy S21", [128, 256], "GB", 3500000),
            ("Samsung Galaxy S21 Plus", [128, 256], "GB", 4000000),
            ("Samsung Galaxy S21 Ultra", [128, 256, 512], "GB", 5500000),
            ("Samsung Galaxy S22", [128, 256, 512], "GB", 4000000),
            ("Samsung Galaxy S22 Plus", [128, 256, 512], "GB", 4500000),
            ("Samsung Galaxy S22 Ultra", [128, 256, 512, 1024], "GB", 6000000),
            ("Samsung Galaxy S23", [128, 256, 512], "GB", 4500000),
            ("Samsung Galaxy S23 Plus", [256, 512], "GB", 5000000),
            ("Samsung Galaxy S23 Ultra", [256, 512, 1024], "GB", 6500000),
            ("Samsung Galaxy S24", [128, 256, 512], "GB", 5500000),
            ("Samsung Galaxy S24 Plus", [256, 512], "GB", 6000000),
            ("Samsung Galaxy S24 Ultra", [256, 512, 1024], "GB", 7000000),
            ("Samsung Galaxy Z Fold 4", [256, 512], "GB", 8500000),
            ("Samsung Galaxy Z Fold 5", [256, 512, 1024], "GB", 9000000),
            ("Samsung Galaxy Z Flip 4", [128, 256], "GB", 5000000),
            ("Samsung Galaxy Z Flip 5", [256, 512], "GB", 5500000),
            ("Samsung Galaxy A54", [128, 256], "GB", 2000000),
            ("Samsung Galaxy A34", [128, 256], "GB", 1800000),
            ("iPad Air 4", [64, 256], "GB", 3000000),
            ("iPad Air 5", [64, 256, 512], "GB", 3500000),
            ("iPad Pro 11 M1", [128, 256, 512, 1024], "GB", 5000000),
            ("iPad Pro 11 M2", [128, 256, 512, 1024], "GB", 5500000),
            ("iPad Pro 12.9 M1", [128, 256, 512, 1024], "GB", 6500000),
            ("iPad Pro 12.9 M2", [256, 512, 1024], "GB", 7000000),
            ("iPad Mini 6", [64, 256], "GB", 3000000),
            ("iPad 9", [64, 256], "GB", 1800000),
            ("iPad 10", [64, 256], "GB", 2500000),
            ("MacBook Air M1", [256, 512], "GB", 4500000),
            ("MacBook Air M2", [256, 512, 1024], "GB", 5500000),
            ("MacBook Air M3", [256, 512, 1024], "GB", 6000000),
            ("MacBook Pro 13 M1", [256, 512], "GB", 6000000),
            ("MacBook Pro 13 M2", [256, 512, 1024], "GB", 7000000),
            ("MacBook Pro 14 M1", [512, 1024], "GB", 8500000),
            ("MacBook Pro 14 M2", [512, 1024], "GB", 9000000),
            ("MacBook Pro 14 M3", [512, 1024], "GB", 9500000),
            ("MacBook Pro 16 M1", [512, 1024], "GB", 10000000),
            ("MacBook Pro 16 M2", [512, 1024], "GB", 10500000),
            ("MacBook Pro 16 M3", [512, 1024], "GB", 11000000),
            ("Apple Watch SE", [40, 44], "mm", 1500000),
            ("Apple Watch Series 8", [41, 45], "mm", 2000000),
            ("Apple Watch Series 9", [41, 45], "mm", 2200000),
            ("Apple Watch Ultra", [49], "mm", 3800000),
            ("Apple Watch Ultra 2", [49], "mm", 4000000),
            ("AirPods 2", [1], "Unidad", 600000),
            ("AirPods 3", [1], "Unidad", 900000),
            ("AirPods Pro", [1], "Unidad", 1000000),
            ("AirPods Pro 2", [1], "Unidad", 1200000),
            ("AirPods Max", [1], "Unidad", 2800000),
            ("Google Pixel 7", [128, 256], "GB", 3000000),
            ("Google Pixel 7 Pro", [128, 256, 512], "GB", 4500000),
            ("Google Pixel 8", [128, 256], "GB", 3500000),
            ("Google Pixel 8 Pro", [128, 256, 512], "GB", 5000000),
            ("Google Pixel Fold", [256, 512], "GB", 9000000),
            ("Xiaomi 12", [128, 256, 512], "GB", 2800000),
            ("Xiaomi 13", [128, 256, 512], "GB", 3000000),
            ("Xiaomi 13 Pro", [256, 512], "GB", 4500000),
            ("Xiaomi 13 Ultra", [256, 512, 1024], "GB", 5500000),
            ("OnePlus 10 Pro", [128, 256], "GB", 3500000),
            ("OnePlus 11", [128, 256, 512], "GB", 3800000),
            ("OnePlus 11 Pro", [256, 512], "GB", 4500000),
            ("Sony Xperia 1 IV", [256, 512], "GB", 5500000),
            ("Sony Xperia 1 V", [256, 512], "GB", 6000000),
            ("Sony Xperia 5 V", [128, 256], "GB", 4500000),
            ("Microsoft Surface Pro 8", [128, 256, 512], "GB", 5000000),
            ("Microsoft Surface Pro 9", [256, 512, 1024], "GB", 5500000),
            ("Microsoft Surface Laptop 4", [256, 512], "GB", 6000000),
            ("Microsoft Surface Laptop 5", [256, 512, 1024], "GB", 6500000),
            ("Huawei P50 Pro", [128, 256, 512], "GB", 4000000),
            ("Huawei Mate 50 Pro", [256, 512], "GB", 5000000),
            ("Motorola Edge 40", [128, 256], "GB", 2500000),
            ("Motorola Edge 40 Pro", [256, 512], "GB", 3500000),
            ("ASUS ROG Phone 7", [256, 512], "GB", 5500000),
            ("ASUS Zenfone 10", [128, 256, 512], "GB", 3500000),
            ("Lenovo Legion Y700", [128, 256], "GB", 2000000),
            ("Nothing Phone 2", [128, 256, 512], "GB", 3000000)
        ]
    }
    
    colors = ["Negro", "Blanco", "Azul", "Rojo", "Plata", "Oro", "Gris Espacial", "Verde", "Morado", "Rosa", 
              "Azul Pacífico", "Verde Alpino", "Titanio Natural", "Titanio Azul", "Titanio Negro", "Titanio Blanco",
              "Medianoche", "Luz Estelar", "Rojo Product", "Amarillo"]

    if color_options is not None:
        colors = list(color_options)

    materials = [
        "Aluminio",
        "Titanio",
        "Vidrio",
        "Cerámica",
    ]

    tech_images = [
        "https://images.unsplash.com/photo-1544244015-0df4b3ffc6b0?auto=format&fit=crop&q=80&w=800", # Laptop/Code
        "https://images.unsplash.com/photo-1496181133206-80ce9b88a853?auto=format&fit=crop&q=80&w=1000", # MacBook

    ]

    def get_random_tech_image():
        return random.choice(tech_images)

    providers = [
        ("TechData Import", 0.75),
        ("iShop Distribution", 0.80),
        ("Global Mobile", 0.70),
        ("Comercializadora Apple", 0.85)
    ]

    # Error types to inject (20 different errors)
    error_types = [
        "duplicate_name",           # 1. Duplicate product name
        "empty_name",               # 2. Empty product name
        "invalid_type",             # 3. Invalid product type
        "missing_category",         # 4. Missing category
        "invalid_unit_group",       # 5. Invalid unit group
        "missing_unit",             # 6. Missing unit
        "empty_description",        # 7. Empty description
        "invalid_perishable",       # 8. Invalid isPerishable value
        "missing_sku",              # 9. Missing SKU
        "duplicate_sku",            # 10. Duplicate SKU
        "invalid_barcode",          # 11. Invalid barcode format
        "negative_cost",            # 12. Negative cost
        "negative_price",           # 13. Negative price
        "zero_price",               # 14. Zero price
        "price_less_than_cost",     # 15. Price less than cost
        "invalid_pum_base",         # 16. Invalid PUM base
        "missing_provider",         # 17. Missing provider name
        "negative_provider_cost",   # 18. Negative provider cost
        "special_chars_name",       # 19. Special characters in name
        "extremely_long_name"       # 20. Extremely long product name
    ]
    
    # Track used names and SKUs for duplicate errors
    used_names = set()
    used_skus = set()
    used_skus_list = []
    duplicate_skus_remaining = int(duplicate_sku_count)
    duplicate_name_target = None
    duplicate_sku_target = None

    product_level_error_types = {
        "duplicate_name",
        "empty_name",
        "invalid_type",
        "missing_category",
        "invalid_unit_group",
        "missing_unit",
        "empty_description",
        "invalid_perishable",
        "invalid_pum_base",
        "missing_provider",
        "special_chars_name",
        "extremely_long_name",
    }

    variant_level_error_types = {
        "missing_sku",
        "duplicate_sku",
        "invalid_barcode",
        "negative_cost",
        "negative_price",
        "zero_price",
        "price_less_than_cost",
        "negative_provider_cost",
    }

    def normalize_type(value):
        v = str(value or '').strip().lower()
        if v in ['product', 'producto', 'productos']:
            return 'Product'
        if v in ['service', 'servicio', 'servicios']:
            return 'Service'
        return value

    products_with_variants = products_with_variants_count
    simple_products = simple_products_count
    
    product_id = 0
    
    # === PART 1: Generate products WITH variants (first 2 products) ===
    for i in range(products_with_variants):
        product_id += 1

        base_name, storage_opts, storage_unit, base_price_ref = random.choice(
            tech_products["Dispositivos móviles y tecnología"]
        )
        storage = random.choice(storage_opts)

        if storage_unit == "Unidad":
            full_name = f"{base_name}"
        else:
            full_name = f"{base_name} {storage}{storage_unit}"
        
        # Ensure unique name
        while full_name in used_names:
            storage = random.choice(storage_opts)
            if storage_unit == "Unidad":
                full_name = f"{base_name} V{product_id}"
            else:
                full_name = f"{base_name} {storage}{storage_unit} V{product_id}"

        cat_name = random.choice(category_options) if category_options else "Dispositivos móviles y tecnología"
        target_group = "Cantidad / Unidades"
        target_unit = "Unidad"

        group_data = next((g for g in units_config if g['name'] == target_group), None)
        group_name = group_data['name'] if group_data else target_group
        unit_name = target_unit

        product_type = 'Productos'
        is_perishable = 'No'
        pum_base = 'Unidad'
        description = f'Descripción original para {full_name}. Garantía 1 año.'

        used_names.add(full_name)

        product_variant_count = int(variants_count_by_product[i])
        blank_material_for_product = int(attribute_count) == 2 and random.random() < float(material_blank_rate)

        for v_i in range(product_variant_count):
            color = colors[v_i % len(colors)]
            material = materials[v_i % len(materials)]
            if blank_material_for_product:
                material = ""
            variation = random.uniform(0.9, 1.1)
            price = int(base_price_ref * variation / 1000) * 1000
            base_price_cost = int(price * 0.7)

            prov_name, prov_margin = random.choice(providers)
            prov_cost = int(price * prov_margin)

            sku = f"{base_name.replace(' ', '')[:4].upper()}-{storage}-{product_id:05d}-{v_i+1:02d}"
            barcode = f'{7700000000000 + (product_id * 10) + v_i}'
            barcode2 = f'{7800000000000 + (product_id * 10) + v_i}'
            barcode3 = f'{7900000000000 + (product_id * 10) + v_i}'

            if duplicate_skus_remaining > 0 and used_skus_list:
                sku = random.choice(used_skus_list)
                duplicate_skus_remaining -= 1

            if inject_errors and random.random() < float(error_rate):
                err = random.choice(
                    [
                        "missing_sku",
                        "invalid_barcode",
                        "negative_cost",
                        "zero_price",
                        "price_less_than_cost",
                        "missing_category",
                        "empty_description",
                        "empty_attr",
                    ]
                )
                if err == "missing_sku":
                    sku = ""
                elif err == "invalid_barcode":
                    barcode = "ABC123"
                elif err == "negative_cost":
                    base_price_cost = -abs(int(base_price_cost))
                elif err == "zero_price":
                    price = 0
                elif err == "price_less_than_cost":
                    price = max(0, int(base_price_cost) - 1)
                elif err == "missing_category":
                    cat_name = ""
                elif err == "empty_description":
                    description = ""
                elif err == "empty_attr":
                    if random.random() < 0.5:
                        color = ""
                    else:
                        material = ""

            used_skus.add(sku)
            if sku:
                used_skus_list.append(sku)

            raw_data = [
                normalize_type(product_type),
                full_name,
                get_random_tech_image(),  # Photo
                cat_name,
                group_name,
                unit_name,
                description,
                is_perishable,
                color,
                *( [material] if int(attribute_count) == 2 else [] ),
                sku,
                barcode,
                barcode2,
                barcode3,
                base_price_cost,
                price,
                pum_base,
                price,
                prov_name,
                prov_cost
            ]

            row_cells = []
            for idx, val in enumerate(raw_data):
                fill = None
                if idx == 4:
                    fill = yellow_fill
                elif idx == 6:
                    fill = blue_fill
                elif idx in (8, 9) and int(attribute_count) == 2:
                    fill = pink_fill
                elif idx == 8 and int(attribute_count) == 1:
                    fill = pink_fill
                row_cells.append(create_styled_cell(val, fill=fill, border=thin_border))
            ws.append(row_cells)
    
    # === PART 2: Generate SIMPLE products (remaining 798 products, 1 row each, no attribute) ===
    for i in range(simple_products):
        product_id += 1

        base_name, storage_opts, storage_unit, base_price_ref = random.choice(
            tech_products["Dispositivos móviles y tecnología"]
        )
        storage = random.choice(storage_opts)

        if storage_unit == "Unidad":
            full_name = f"{base_name} Simple {product_id}"
        else:
            full_name = f"{base_name} {storage}{storage_unit} Simple {product_id}"
        
        # Ensure unique name
        while full_name in used_names:
            full_name = f"{base_name} Simple {product_id}-{random.randint(1000, 9999)}"

        cat_name = random.choice(category_options) if category_options else "Dispositivos móviles y tecnología"
        target_group = "Cantidad / Unidades"
        target_unit = "Unidad"

        group_data = next((g for g in units_config if g['name'] == target_group), None)
        group_name = group_data['name'] if group_data else target_group
        unit_name = target_unit

        product_type = 'Productos'
        is_perishable = 'No'
        pum_base = 'Unidad'
        description = f'Descripción para {full_name}. Garantía 1 año.'

        used_names.add(full_name)

        # Single row - no attributes
        variation = random.uniform(0.9, 1.1)
        price = int(base_price_ref * variation / 1000) * 1000
        base_price_cost = int(price * 0.7)

        prov_name, prov_margin = random.choice(providers)
        prov_cost = int(price * prov_margin)

        sku = f"SIMP-{product_id:05d}"
        barcode = f'{7700000000000 + (product_id * 10)}'
        barcode2 = f'{7800000000000 + (product_id * 10)}'
        barcode3 = f'{7900000000000 + (product_id * 10)}'

        if duplicate_skus_remaining > 0 and used_skus_list:
            sku = random.choice(used_skus_list)
            duplicate_skus_remaining -= 1

        used_skus.add(sku)
        if sku:
            used_skus_list.append(sku)

        raw_data = [
            normalize_type(product_type),
            full_name,
            get_random_tech_image(),  # Photo
            cat_name,
            group_name,
            unit_name,
            description,
            is_perishable,
            '',
            *( [''] if int(attribute_count) == 2 else [] ),
            sku,
            barcode,
            barcode2,
            barcode3,
            base_price_cost,
            price,
            pum_base,
            price,
            prov_name,
            prov_cost
        ]

        row_cells = []
        for idx, val in enumerate(raw_data):
            fill = None
            if idx == 4:
                fill = yellow_fill
            elif idx == 6:
                fill = blue_fill
            row_cells.append(create_styled_cell(val, fill=fill, border=thin_border))
        ws.append(row_cells)

    wb.save(file_path)
    print(f"Successfully created {filename} with {product_count} products and {total_rows} rows.")

if __name__ == "__main__":
    presets = {
        "1000_mixed_ok": {
            "filename": "tech_products_1000_mixed_ok.xlsx",
            "product_count": 100,
            "variants_per_product": 5,
            "products_with_variants": 2,
            "attribute_count": 1,
            "inject_errors": False,
            "error_rate": 0.0,
            "seed": 42,
        },
        "1000_mixed_err1": {
            "filename": "tech_products_1000_mixed_err1.xlsx",
            "product_count": 1000,
            "variants_per_product": 2,
            "products_with_variants": 2,
            "attribute_count": 2,
            "inject_errors": True,
            "error_rate": 0.01,
            "seed": 42,
        },
        "5000_mixed_err1": {
            "filename": "tech_products_5000_mixed_err1.xlsx",
            "product_count": 5000,
            "variants_per_product": 2,
            "products_with_variants": 2,
            "attribute_count": 2,
            "inject_errors": True,
            "error_rate": 0.01,
            "seed": 42,
        },
        "10000_mixed_ok": {
            "filename": "tech_products_10000_mixed_ok.xlsx",
            "product_count": 10000,
            "variants_per_product": 2,
            "products_with_variants": 2,
            "attribute_count": 2,
            "inject_errors": False,
            "error_rate": 0.0,
            "seed": 42,
        },
        "10000_mixed_err5": {
            "filename": "tech_products_10000_mixed_err5.xlsx",
            "product_count": 10000,
            "variants_per_product": 2,
            "products_with_variants": 2,
            "attribute_count": 2,
            "inject_errors": True,
            "error_rate": 0.05,
            "seed": 42,
        },
        "1000_2props_ok": {
            "filename": "tech_products_1000_2props_ok.xlsx",
            "product_count": 1000,
            "variants_per_product": 2,
            "products_with_variants": 1000,
            "attribute_count": 2,
            "inject_errors": False,
            "error_rate": 0.0,
            "seed": 42,
        },
        "1000_2props_err_random": {
            "filename": "tech_products_1000_2props_err_random.xlsx",
            "product_count": 1000,
            "variants_per_product": 2,
            "products_with_variants": 1000,
            "attribute_count": 2,
            "inject_errors": True,
            "error_rate": 0.02,
            "seed": 42,
        },
        "1000_1prop_ok": {
            "filename": "tech_products_1000_1prop_ok.xlsx",
            "product_count": 1000,
            "variants_per_product": 2,
            "products_with_variants": 1000,
            "attribute_count": 1,
            "inject_errors": False,
            "error_rate": 0.0,
            "seed": 42,
        },
        "1000_1prop_err_random": {
            "filename": "tech_products_1000_1prop_err_random.xlsx",
            "product_count": 1000,
            "variants_per_product": 2,
            "products_with_variants": 1000,
            "attribute_count": 1,
            "inject_errors": True,
            "error_rate": 0.02,
            "seed": 42,
        },
    }

    parser = argparse.ArgumentParser()
    parser.add_argument("--preset", type=str, choices=sorted(presets.keys()))
    parser.add_argument("--filename", type=str)
    parser.add_argument("--product-count", type=int)
    parser.add_argument("--variants-per-product", type=int)
    parser.add_argument("--variants-per-product-mix", type=str)
    parser.add_argument("--products-with-variants", type=int)
    parser.add_argument("--attribute-count", type=int, choices=[1, 2])
    parser.add_argument("--material-blank-rate", type=float)
    parser.add_argument("--color-options", type=str)
    parser.add_argument("--duplicate-sku-count", type=int)
    parser.add_argument("--category-options", type=str)
    parser.add_argument("--inject-errors", action="store_true")
    parser.add_argument("--no-inject-errors", action="store_true")
    parser.add_argument("--error-rate", type=float)
    parser.add_argument("--seed", type=int)

    args = parser.parse_args()

    if not args.preset and not any(
        v is not None
        for v in [
            args.filename,
            args.product_count,
            args.variants_per_product,
            args.error_rate,
            args.seed,
        ]
    ):
        parser.print_help()
        print("\nPresets:")
        for k in sorted(presets.keys()):
            print(f"  {k}")
        raise SystemExit(0)

    config = dict(presets.get(args.preset, {}))

    if args.filename is not None:
        config["filename"] = args.filename
    if args.product_count is not None:
        config["product_count"] = args.product_count
    if args.variants_per_product is not None:
        config["variants_per_product"] = args.variants_per_product
    if args.variants_per_product_mix is not None:
        config["variants_per_product_mix"] = args.variants_per_product_mix
    if args.products_with_variants is not None:
        config["products_with_variants"] = args.products_with_variants
    if args.attribute_count is not None:
        config["attribute_count"] = args.attribute_count
    if args.material_blank_rate is not None:
        config["material_blank_rate"] = args.material_blank_rate
    if args.color_options is not None:
        config["color_options"] = args.color_options
    if args.duplicate_sku_count is not None:
        config["duplicate_sku_count"] = args.duplicate_sku_count
    if args.category_options is not None:
        config["category_options"] = args.category_options
    if args.error_rate is not None:
        config["error_rate"] = args.error_rate
    if args.seed is not None:
        config["seed"] = args.seed

    if args.no_inject_errors:
        config["inject_errors"] = False
    elif args.inject_errors:
        config["inject_errors"] = True

    required = ["filename", "product_count"]
    missing = [k for k in required if config.get(k) in (None, "")]
    if missing:
        raise SystemExit(f"Missing required args: {', '.join(missing)}")

    variants_mix = config.get("variants_per_product_mix")
    if isinstance(variants_mix, str) and variants_mix.strip():
        variants_mix = [p.strip() for p in variants_mix.split(",") if p.strip()]
        variants_mix = [int(p) for p in variants_mix]
    else:
        variants_mix = None

    color_options = config.get("color_options")
    if isinstance(color_options, str) and color_options.strip():
        color_options = [p.strip() for p in color_options.split(",") if p.strip()]
    else:
        color_options = None

    category_options = config.get("category_options")
    if isinstance(category_options, str) and category_options.strip():
        category_options = [p.strip() for p in category_options.split(",") if p.strip()]
    else:
        category_options = None

    generate_data(
        config["filename"],
        product_count=config["product_count"],
        variants_per_product=config.get("variants_per_product", 2),
        products_with_variants=config.get("products_with_variants"),
        attribute_count=int(config.get("attribute_count", 2)),
        variants_per_product_mix=variants_mix,
        material_blank_rate=float(config.get("material_blank_rate", 0.0)),
        color_options=color_options,
        duplicate_sku_count=int(config.get("duplicate_sku_count", 0) or 0),
        category_options=category_options,
        inject_errors=bool(config.get("inject_errors", False)),
        error_rate=float(config.get("error_rate", 0.0)),
        seed=config.get("seed"),
    )
